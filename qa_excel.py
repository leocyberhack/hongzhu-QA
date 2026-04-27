from __future__ import annotations

from collections import OrderedDict
from copy import copy
from dataclasses import dataclass, field
from pathlib import Path
import os
import stat
import uuid

import openpyxl


HEADERS = ["分类", "区域", "POI", "咨询意图", "具体问法", "回复内容"]
META_HEADER = "__QA_MANAGER_TYPE"
TEMPLATE_RECORD = "template"
CUSTOM_RECORD = "custom"
SUPPORTED_RECORD_TYPES = {TEMPLATE_RECORD, CUSTOM_RECORD}


def is_blank(value: object) -> bool:
    return value is None or (isinstance(value, str) and value.strip() == "")


def clean(value: object) -> str:
    if value is None:
        return ""
    return str(value).replace("\n", " ").strip()


@dataclass
class SheetStandardizationReport:
    sheet: str
    merged_a_to_d_count: int = 0
    rows_missing_a_to_d: int = 0
    qa_rows: int = 0
    issues: list[str] = field(default_factory=list)


@dataclass
class StandardizationReport:
    complete: bool
    sheets: list[SheetStandardizationReport]

    @property
    def issue_count(self) -> int:
        return sum(len(sheet.issues) for sheet in self.sheets)


@dataclass
class QARecord:
    sheet: str
    category: str
    region: str
    poi: str
    intent: str
    question: str
    answer: str
    row: int | None = None
    record_type: str = TEMPLATE_RECORD
    record_id: str = field(default_factory=lambda: uuid.uuid4().hex)


@dataclass
class TemplateItem:
    intent: str
    question: str
    source_intent: str | None = None
    source_question: str | None = None

    @property
    def key(self) -> tuple[str, str]:
        return self.intent, self.question

    @property
    def source_key(self) -> tuple[str, str] | None:
        if self.source_intent is None or self.source_question is None:
            return None
        return self.source_intent, self.source_question


class QAWorkbook:
    def __init__(self, path: str | Path | None = None):
        self.path: Path | None = Path(path) if path else None
        self.records: list[QARecord] = []
        self.templates: OrderedDict[str, list[TemplateItem]] = OrderedDict()
        self.sheet_order: list[str] = []
        self.category_sheet: OrderedDict[str, str] = OrderedDict()
        self.report: StandardizationReport | None = None

    def load(self, path: str | Path | None = None) -> None:
        if path is not None:
            self.path = Path(path)
        if self.path is None:
            raise ValueError("未选择 Excel 文件")
        if not self.path.exists():
            raise FileNotFoundError(self.path)

        workbook = openpyxl.load_workbook(self.path, data_only=False)
        self.report = scan_standardization(self.path)
        self.records = []
        self.sheet_order = [sheet.title for sheet in workbook.worksheets]

        for worksheet in workbook.worksheets:
            if worksheet.max_row < 2:
                continue

            has_meta_column = clean(worksheet.cell(1, 7).value) == META_HEADER
            current = ["", "", "", ""]
            for row_idx in range(2, worksheet.max_row + 1):
                values = [clean(worksheet.cell(row_idx, col).value) for col in range(1, 7)]
                record_type = clean(worksheet.cell(row_idx, 7).value) if has_meta_column else ""
                if record_type not in SUPPORTED_RECORD_TYPES:
                    record_type = ""

                for idx in range(4):
                    if values[idx]:
                        current[idx] = values[idx]
                    else:
                        values[idx] = current[idx]

                question = values[4]
                answer = values[5]
                if not question and not answer:
                    continue

                record = QARecord(
                    sheet=worksheet.title,
                    category=values[0],
                    region=values[1],
                    poi=values[2],
                    intent=values[3],
                    question=question,
                    answer=answer,
                    row=row_idx,
                    record_type=record_type,
                )
                self.records.append(record)

                if record.category and record.category not in self.category_sheet:
                    self.category_sheet[record.category] = record.sheet

        self._normalize_record_types()
        self._rebuild_templates()

    @property
    def standardized(self) -> bool:
        return bool(self.report and self.report.complete)

    def standardize_in_place(self) -> None:
        if self.path is None:
            raise ValueError("未选择 Excel 文件")
        standardize_workbook_in_place(self.path)
        self.load(self.path)

    def save_all(self) -> None:
        if self.path is None:
            raise ValueError("未选择 Excel 文件")
        save_records_to_workbook(self.path, self.records, self.sheet_order)
        self.load(self.path)

    def categories(self) -> list[str]:
        return list(OrderedDict((record.category, None) for record in self.records if record.category).keys())

    def regions(self, category: str) -> list[str]:
        return list(
            OrderedDict(
                (record.region, None)
                for record in self.records
                if record.category == category and record.region
            ).keys()
        )

    def pois(self, category: str, region: str) -> list[str]:
        return list(
            OrderedDict(
                (record.poi, None)
                for record in self.records
                if record.category == category and record.region == region and record.poi
            ).keys()
        )

    def intents(self, category: str, region: str, poi: str) -> list[str]:
        return list(
            OrderedDict(
                (record.intent, None)
                for record in self.records
                if record.category == category
                and record.region == region
                and record.poi == poi
                and record.intent
            ).keys()
        )

    def qa_pairs(self, category: str, region: str, poi: str, intent: str) -> list[QARecord]:
        return [
            record
            for record in self.records
            if record.category == category
            and record.region == region
            and record.poi == poi
            and record.intent == intent
        ]

    def matching_question_records(self, category: str, intent: str, question: str) -> list[QARecord]:
        return [
            record
            for record in self.records
            if record.category == category and record.intent == intent and record.question == question
        ]

    def bulk_update_answer(self, category: str, intent: str, question: str, answer: str) -> tuple[int, int]:
        matches = self.matching_question_records(category, intent, question)
        if not matches:
            raise ValueError("没有找到可批量修改的匹配问答")
        for record in matches:
            record.answer = answer
        poi_count = len({(record.region, record.poi) for record in matches})
        self.save_all()
        return len(matches), poi_count

    def template_for_category(self, category: str) -> list[TemplateItem]:
        return [
            TemplateItem(item.intent, item.question, item.intent, item.question)
            for item in self.templates.get(category, [])
        ]

    def add_poi(self, category: str, region: str, poi: str) -> None:
        category = category.strip()
        region = region.strip()
        poi = poi.strip()
        if not category or not region or not poi:
            raise ValueError("分类、区域和 POI 都不能为空")
        if region not in self.regions(category):
            raise ValueError(f"区域“{region}”不属于分类“{category}”，请从该分类已有区域中选择")
        if category not in self.templates or not self.templates[category]:
            raise ValueError(f"分类“{category}”还没有可套用的问题模板")
        if any(
            record.category == category and record.region == region and record.poi == poi
            for record in self.records
        ):
            raise ValueError(f"“{region} / {poi}”已经存在")

        sheet = self.category_sheet.get(category)
        if not sheet:
            raise ValueError(f"找不到分类“{category}”对应的工作表")

        for item in self.templates[category]:
            self.records.append(
                QARecord(
                    sheet=sheet,
                    category=category,
                    region=region,
                    poi=poi,
                    intent=item.intent,
                    question=item.question,
                    answer="",
                    record_type=TEMPLATE_RECORD,
                )
            )
        self.save_all()

    def add_qa_pair(self, category: str, region: str, poi: str, intent: str, question: str, answer: str) -> None:
        category = category.strip()
        region = region.strip()
        poi = poi.strip()
        intent = intent.strip()
        question = question.strip()
        if not category or not region or not poi or not intent or not question:
            raise ValueError("分类、区域、POI、咨询意图和具体问法都不能为空")

        sheet = ""
        for record in self.records:
            if record.category == category and record.region == region and record.poi == poi:
                sheet = record.sheet
                break
        if not sheet:
            sheet = self.category_sheet.get(category, "")
        if not sheet:
            raise ValueError(f"找不到分类“{category}”对应的工作表")

        duplicate = any(
            record.category == category
            and record.region == region
            and record.poi == poi
            and record.intent == intent
            and record.question == question
            for record in self.records
        )
        if duplicate:
            raise ValueError("该 POI 的当前意图下已经存在相同问法")

        self.records.append(
            QARecord(
                sheet=sheet,
                category=category,
                region=region,
                poi=poi,
                intent=intent,
                question=question,
                answer=answer,
                record_type=CUSTOM_RECORD,
            )
        )
        self.save_all()

    def delete_poi(self, category: str, region: str, poi: str) -> int:
        before = len(self.records)
        self.records = [
            record
            for record in self.records
            if not (record.category == category and record.region == region and record.poi == poi)
        ]
        removed = before - len(self.records)
        if removed:
            self.save_all()
        return removed

    def rename_poi(self, category: str, region: str, old_poi: str, new_poi: str) -> int:
        new_poi = new_poi.strip()
        if not new_poi:
            raise ValueError("新的 POI 名称不能为空")
        if any(
            record.category == category
            and record.region == region
            and record.poi == new_poi
            for record in self.records
        ):
            raise ValueError(f"“{region} / {new_poi}”已经存在")

        changed = 0
        for record in self.records:
            if record.category == category and record.region == region and record.poi == old_poi:
                record.poi = new_poi
                changed += 1
        if changed:
            self.save_all()
        return changed

    def apply_template(self, category: str, template_items: list[TemplateItem]) -> None:
        if not category:
            raise ValueError("请选择分类")
        normalized = [
            TemplateItem(item.intent.strip(), item.question.strip(), item.source_intent, item.source_question)
            for item in template_items
            if item.intent.strip() or item.question.strip()
        ]
        if not normalized:
            raise ValueError("模板不能为空")

        seen: set[tuple[str, str]] = set()
        for item in normalized:
            if not item.intent or not item.question:
                raise ValueError("模板里的咨询意图和具体问法都不能为空")
            if item.key in seen:
                raise ValueError(f"模板存在重复问法：{item.intent} / {item.question}")
            seen.add(item.key)

        old_template_keys = {item.key for item in self.templates.get(category, [])}
        poi_order: OrderedDict[tuple[str, str, str, str], None] = OrderedDict()
        answer_map: dict[tuple[str, str, str, str], dict[tuple[str, str], str]] = {}
        custom_records: dict[tuple[str, str, str, str], list[QARecord]] = {}
        for record in self.records:
            if record.category != category:
                continue
            poi_key = (record.sheet, record.category, record.region, record.poi)
            poi_order.setdefault(poi_key, None)
            answer_map.setdefault(poi_key, {})[(record.intent, record.question)] = record.answer
            if record.record_type == CUSTOM_RECORD or (record.intent, record.question) not in old_template_keys:
                record.record_type = CUSTOM_RECORD
                custom_records.setdefault(poi_key, []).append(record)

        rebuilt_pois: set[tuple[str, str, str, str]] = set()
        new_records: list[QARecord] = []
        for record in self.records:
            if record.category != category:
                new_records.append(record)
                continue

            poi_key = (record.sheet, record.category, record.region, record.poi)
            if poi_key in rebuilt_pois:
                continue

            existing_answers = answer_map.get(poi_key, {})
            for item in normalized:
                answer = ""
                if item.key in existing_answers:
                    answer = existing_answers[item.key]
                elif item.source_key and item.source_key in existing_answers:
                    answer = existing_answers[item.source_key]

                new_records.append(
                    QARecord(
                        sheet=poi_key[0],
                        category=poi_key[1],
                        region=poi_key[2],
                        poi=poi_key[3],
                        intent=item.intent,
                        question=item.question,
                        answer=answer,
                        record_type=TEMPLATE_RECORD,
                    )
                )
            new_records.extend(custom_records.get(poi_key, []))
            rebuilt_pois.add(poi_key)

        self.records = new_records
        self.save_all()

    def _normalize_record_types(self) -> None:
        category_pois: OrderedDict[str, OrderedDict[tuple[str, str, str], None]] = OrderedDict()
        pair_pois: OrderedDict[str, OrderedDict[tuple[str, str], set[tuple[str, str, str]]]] = OrderedDict()

        for record in self.records:
            if not record.category or not record.intent or not record.question:
                continue
            poi_key = (record.sheet, record.region, record.poi)
            key = (record.intent, record.question)
            category_pois.setdefault(record.category, OrderedDict()).setdefault(poi_key, None)
            pair_pois.setdefault(record.category, OrderedDict()).setdefault(key, set()).add(poi_key)

        for record in self.records:
            if record.record_type in SUPPORTED_RECORD_TYPES:
                continue
            poi_count = len(category_pois.get(record.category, {}))
            key = (record.intent, record.question)
            key_poi_count = len(pair_pois.get(record.category, {}).get(key, set()))
            record.record_type = TEMPLATE_RECORD if poi_count and key_poi_count == poi_count else CUSTOM_RECORD

    def _rebuild_templates(self) -> None:
        category_pois: OrderedDict[str, OrderedDict[tuple[str, str, str], None]] = OrderedDict()
        pair_pois: OrderedDict[str, OrderedDict[tuple[str, str], set[tuple[str, str, str]]]] = OrderedDict()
        pair_order: OrderedDict[str, OrderedDict[tuple[str, str], TemplateItem]] = OrderedDict()

        for record in self.records:
            if not record.category or not record.intent or not record.question:
                continue
            poi_key = (record.sheet, record.region, record.poi)
            category_pois.setdefault(record.category, OrderedDict()).setdefault(poi_key, None)
            if record.record_type == CUSTOM_RECORD:
                continue
            key = (record.intent, record.question)
            pair_pois.setdefault(record.category, OrderedDict()).setdefault(key, set()).add(poi_key)
            pair_order.setdefault(record.category, OrderedDict()).setdefault(
                key,
                TemplateItem(record.intent, record.question, record.intent, record.question),
            )

        templates: OrderedDict[str, list[TemplateItem]] = OrderedDict()
        for category, ordered_items in pair_order.items():
            poi_count = len(category_pois.get(category, {}))
            templates[category] = [
                item
                for key, item in ordered_items.items()
                if poi_count and len(pair_pois[category].get(key, set())) == poi_count
            ]
        self.templates = templates


def scan_standardization(path: str | Path) -> StandardizationReport:
    workbook = openpyxl.load_workbook(path, data_only=False)
    reports: list[SheetStandardizationReport] = []

    for worksheet in workbook.worksheets:
        sheet_report = SheetStandardizationReport(sheet=worksheet.title)

        for merged in worksheet.merged_cells.ranges:
            min_col, min_row, max_col, max_row = merged.bounds
            intersects_a_to_d = min_col <= 4 and max_col >= 1
            touches_data_rows = max_row >= 2
            if intersects_a_to_d and touches_data_rows:
                sheet_report.merged_a_to_d_count += 1

        for row_idx in range(2, worksheet.max_row + 1):
            has_qa = any(not is_blank(worksheet.cell(row_idx, col).value) for col in (5, 6))
            if not has_qa:
                continue
            sheet_report.qa_rows += 1
            if any(is_blank(worksheet.cell(row_idx, col).value) for col in range(1, 5)):
                sheet_report.rows_missing_a_to_d += 1

        if sheet_report.merged_a_to_d_count:
            sheet_report.issues.append(f"A-D 列仍有 {sheet_report.merged_a_to_d_count} 个合并单元格")
        if sheet_report.rows_missing_a_to_d:
            sheet_report.issues.append(f"有 {sheet_report.rows_missing_a_to_d} 条 QA 的 A-D 列不完整")
        reports.append(sheet_report)

    complete = all(not report.issues for report in reports)
    return StandardizationReport(complete=complete, sheets=reports)


def standardize_workbook_in_place(path: str | Path) -> None:
    path = Path(path)
    workbook = openpyxl.load_workbook(path, data_only=False)

    for worksheet in workbook.worksheets:
        merged_ranges = list(worksheet.merged_cells.ranges)
        for merged in merged_ranges:
            min_col, min_row, max_col, max_row = merged.bounds
            if min_col > 4 or max_col < 1 or max_row < 2:
                continue

            top_left = worksheet.cell(min_row, min_col)
            value = top_left.value
            style = copy(top_left._style)
            font = copy(top_left.font)
            fill = copy(top_left.fill)
            border = copy(top_left.border)
            alignment = copy(top_left.alignment)
            protection = copy(top_left.protection)
            number_format = top_left.number_format

            worksheet.unmerge_cells(str(merged))
            for row_idx in range(min_row, max_row + 1):
                for col_idx in range(min_col, max_col + 1):
                    if col_idx > 4:
                        continue
                    cell = worksheet.cell(row_idx, col_idx)
                    cell.value = value
                    cell._style = copy(style)
                    cell.font = copy(font)
                    cell.fill = copy(fill)
                    cell.border = copy(border)
                    cell.alignment = copy(alignment)
                    cell.protection = copy(protection)
                    cell.number_format = number_format

        last_values = ["", "", "", ""]
        for row_idx in range(2, worksheet.max_row + 1):
            has_qa = any(not is_blank(worksheet.cell(row_idx, col).value) for col in (5, 6))
            if not has_qa:
                continue
            for col_idx in range(1, 5):
                cell = worksheet.cell(row_idx, col_idx)
                if is_blank(cell.value):
                    if last_values[col_idx - 1]:
                        cell.value = last_values[col_idx - 1]
                else:
                    last_values[col_idx - 1] = cell.value

    save_workbook(path, workbook)


def save_records_to_workbook(path: str | Path, records: list[QARecord], sheet_order: list[str]) -> None:
    path = Path(path)
    workbook = openpyxl.load_workbook(path, data_only=False)
    records_by_sheet: OrderedDict[str, list[QARecord]] = OrderedDict()

    known_sheets = sheet_order or [sheet.title for sheet in workbook.worksheets]
    for sheet_name in known_sheets:
        records_by_sheet[sheet_name] = []
    for record in records:
        records_by_sheet.setdefault(record.sheet, []).append(record)

    for sheet_name, sheet_records in records_by_sheet.items():
        if sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
        else:
            worksheet = workbook.create_sheet(sheet_name)

        for col_idx, header in enumerate(HEADERS, 1):
            worksheet.cell(1, col_idx).value = header
        worksheet.cell(1, 7).value = META_HEADER
        worksheet.column_dimensions["G"].hidden = True

        style_source_row = 2 if worksheet.max_row >= 2 else 1
        style_templates = [_snapshot_style(worksheet.cell(style_source_row, col_idx)) for col_idx in range(1, 7)]
        meta_style = _snapshot_style(worksheet.cell(style_source_row, 7))

        if worksheet.max_row >= 2:
            worksheet.delete_rows(2, worksheet.max_row - 1)

        for record in sheet_records:
            row_idx = worksheet.max_row + 1
            record_type = record.record_type if record.record_type in SUPPORTED_RECORD_TYPES else TEMPLATE_RECORD
            values = [record.category, record.region, record.poi, record.intent, record.question, record.answer]
            for col_idx, value in enumerate(values, 1):
                cell = worksheet.cell(row_idx, col_idx)
                cell.value = value
                _apply_style(cell, style_templates[col_idx - 1])
            meta_cell = worksheet.cell(row_idx, 7)
            meta_cell.value = record_type
            _apply_style(meta_cell, meta_style)

    save_workbook(path, workbook)


def save_workbook(path: Path, workbook) -> None:
    _make_writable(path)
    workbook.save(path)


def _make_writable(path: Path) -> None:
    if not path.exists():
        return
    try:
        os.chmod(path, stat.S_IREAD | stat.S_IWRITE)
    except OSError:
        pass


def _snapshot_style(cell) -> dict[str, object]:
    return {
        "style": copy(cell._style),
        "font": copy(cell.font),
        "fill": copy(cell.fill),
        "border": copy(cell.border),
        "alignment": copy(cell.alignment),
        "protection": copy(cell.protection),
        "number_format": cell.number_format,
    }


def _apply_style(cell, snapshot: dict[str, object]) -> None:
    cell._style = copy(snapshot["style"])
    cell.font = copy(snapshot["font"])
    cell.fill = copy(snapshot["fill"])
    cell.border = copy(snapshot["border"])
    cell.alignment = copy(snapshot["alignment"])
    cell.protection = copy(snapshot["protection"])
    cell.number_format = snapshot["number_format"]
